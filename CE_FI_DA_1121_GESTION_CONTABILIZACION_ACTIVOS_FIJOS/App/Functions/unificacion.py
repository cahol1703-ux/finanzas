import os
import glob
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from typing import Dict, Optional, cast
from .logs_config import configurar_logger
from .descarga_informes import verficacion_carpetas
logger = configurar_logger()


def obtecion_de_archivos(directorio: str, carpetas: list[str]) -> Dict[str, Optional[str]]:  # type: ignore[type-arg]
    resultados = {}
    for carpeta in carpetas:
        ruta_completa = os.path.join(directorio, carpeta)
        # Verifica si la ruta existe y es una carpeta
        if os.path.exists(ruta_completa) and os.path.isdir(ruta_completa):
            archivo_xls = glob.glob(os.path.join(ruta_completa, '*xls')) # Busca archivos .xls
            if archivo_xls:
                # Selecciona el archivo más reciente
                archivo = max(archivo_xls, key=os.path.getmtime) # Usa el más reciente
                resultados[carpeta] = archivo
            else:
                resultados[carpeta] = None
        else:
            resultados[carpeta] = None
    return resultados

def leer_archivos_xml(directorio: str, carpetas: list[str]) -> Dict[str, pd.DataFrame]:  # type: ignore[type-arg]
    try:
        archivos_dict = obtecion_de_archivos(directorio, carpetas)
    except Exception as e:
        logger.error(f"Error al cargar archivos: {e}")
        return {}

    dataframes = {}

    for carpeta, archivo in archivos_dict.items():
        # Si no se encuentra archivo para la carpeta, lo ignora
        if not archivo:
            logger.error(f"No se encontró un archivo para la carpeta '{carpeta}'")
            continue

        try:
            # Primer intento: parsear normalmente
            tree = ET.parse(archivo)
            root = tree.getroot()
        except ET.ParseError:
            # Si falla, intenta reparar caracteres inválidos 
            logger.warning(f"Intentando reparar el archivo XML mal formado: {archivo}")
            try:
                with open(archivo, 'r', encoding='utf-8', errors='ignore') as f:
                    contenido = f.read()

                # Reparar caracteres conflictivos
                contenido = contenido.replace('&', '&amp;')

                # Volver a intentar con el contenido reparado
                root = ET.fromstring(contenido)
            except Exception as e:
                logger.error(f"❌ No se pudo reparar ni leer el archivo {archivo}: {e}")
                continue

        try:
            # Procesar el contenido XML
            filas = []
            max_columnas = 0
            ns = '{urn:schemas-microsoft-com:office:spreadsheet}' # Espacio de nombres usado en XML

            for row_idx, row in enumerate(root.iter(f'{ns}Row')):
                celdas = []
                col_idx = 0
                for cell in row.iter(f'{ns}Cell'):
                    index = cell.attrib.get(f'{ns}Index')
                    if index:
                        index = int(index) - 1
                        # Inserta valores vacíos si faltan columnas intermedias
                        while col_idx < index:
                            celdas.append(None)
                            col_idx += 1
                    data = cell.find(f'{ns}Data')
                    celdas.append(data.text if data is not None else None)
                    col_idx += 1
                # Determina la cantidad máxima de columnas
                if row_idx == 0:
                    max_columnas = len(celdas)
                while len(celdas) < max_columnas:
                    celdas.append(None)

                filas.append(celdas)
            # Convierte a DataFrame y asigna la primera fila como encabezados
            df = pd.DataFrame(filas)
            df.columns = df.iloc[0]
            df = df[1:]
            df.reset_index(drop=True, inplace=True)
            dataframes[carpeta] = df
        except Exception as e:
            logger.error(f"Error al procesar el archivo {archivo}: {e}")
    return dataframes

def exportar_excel(directorio, nombre, carpetas):
    try:
        dict_dfs = leer_archivos_xml(directorio, carpetas)
    except Exception as e:
        logger.error(f"error en leer los archivos:  {e}")
        return None

    # CORRECCIÓN 2: Validar que se leyó al menos un DataFrame antes de continuar.
    # Antes, si dict_dfs quedaba vacío (ej. todas las compañías fallaron), el pd.concat
    # lanzaba una excepción no controlada y el archivo nunca se generaba.
    if not dict_dfs:
        logger.error("No se encontró ningún DataFrame para unificar. No se generará el archivo Excel.")
        return None

    # Advertir si faltan compañías respecto a las esperadas
    companias_faltantes = [c for c in carpetas if c not in dict_dfs]
    if companias_faltantes:
        logger.warning(
            f"El Excel se generará incompleto. No se encontraron datos para las compañías: {companias_faltantes}"
        )

    os.makedirs(directorio, exist_ok=True)
    ruta_salida = os.path.join(directorio, nombre)
    # Si ya existe un archivo y se han verificado las carpetas, no se sobreescribe
    if os.path.exists(ruta_salida) and verficacion_carpetas(directorio, carpetas):
        logger.info(f"El archivo {nombre} ya existe. No se guardará el archivo para evitar sobrescribir.")
        return ruta_salida

    for clave, df in dict_dfs.items():
        df.insert(0, "Responsable", "")
        if df.shape[1] > 6:
            try:
                df.iloc[:, 6] = pd.to_datetime(df.iloc[:, 6], errors='coerce').dt.date
            except Exception as e:
                logger.error(f"Error al convertir fechas en {clave}: {e}")
        # Redondear la columna D (índice 3) a 2 decimales si es numérica
        if df.shape[1] > 3:
            try:
                numeric_col = pd.to_numeric(df.iloc[:, 3], errors='coerce')
                df.iloc[:, 3] = numeric_col.round(2)  # type: ignore[reportUnknownMemberType]
            except Exception as e:
                logger.error(f"Error al redondear columna D en {clave}: {e}")
        for idx in [9, 13]:
            if df.shape[1] > idx:
                try:
                    numeric_col = pd.to_numeric(df.iloc[:, idx], errors='coerce')
                    df.iloc[:, idx] = numeric_col.fillna(0).astype(int)  # type: ignore[reportUnknownMemberType]
                except Exception as e:
                    logger.error(f"Error al procesar columna {idx} sin decimales en {clave}: {e}")
        if df.shape[1] > 4:
            try:
                col_4_processed = cast(pd.Series, df.iloc[:, 4].apply(lambda x: str(x).strip() if x is not None else x))
                df.iloc[:, 4] = col_4_processed
                mask = (df.iloc[:, 4] != '2A') & df.iloc[:, 4].notna()
                df = df[mask]
                dict_dfs[clave] = df
            except Exception as e:
                logger.error(f"Error al procesar columna registros 2A ni vacios en {clave}: {e}")

    # CORRECCIÓN 2 (continuación): Verificar de nuevo que queden DataFrames válidos
    # tras el filtrado (en caso extremo de que todos queden vacíos tras eliminar 2A).
    dfs_validos = [df for df in dict_dfs.values() if df is not None and not df.empty]
    if not dfs_validos:
        logger.error("Todos los DataFrames quedaron vacíos tras el filtrado. No se generará el archivo Excel.")
        return None

    df_total = pd.concat(dfs_validos, ignore_index=True)
    df_total.to_excel(ruta_salida, index=False, sheet_name="Transac. Pend. Por Contabilizar")
    # Cargar el archivo con openpyxl
    wb = load_workbook(ruta_salida)
    ws = wb.active
    if ws is None:
        logger.error("No se pudo acceder a la hoja activa del libro Excel")
        return

    # Definir rango de la tabla (A1 hasta la última columna y fila con datos)
    max_col = ws.max_column
    max_row = ws.max_row
    col_letter = ws.cell(row=1, column=max_col).column_letter  # type: ignore[reportAttributeAccessIssue]
    tabla_rango = f"A1:{col_letter}{max_row}"

    # Crear e insertar la tabla
    tabla = Table(displayName="TablaUnificada", ref=tabla_rango)
    estilo = TableStyleInfo(
        name="TableStyleMedium11",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False)
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)
    for idx, cell in enumerate(ws[1], 1):  # ws[1] = primera fila
        if str(cell.value).strip().lower() in ["fecha libro", "fecha libro mayor"]:
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=idx, max_col=idx):
                for celda in row:
                    celda.number_format = 'yyyy/mm/dd'
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # type: ignore[reportAttributeAccessIssue]
        for cell in col:
            try:
                valor = str(cell.value)
                if len(valor) > max_length:
                    max_length = len(valor)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Guardar los cambios
    wb.save(ruta_salida)
    logger.info(f"archivo guardado correctamente en: {ruta_salida}")
    return ruta_salida

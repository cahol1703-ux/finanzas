import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from .logs_config import configurar_logger
from .hojas_excel import obtener_hoja_excel

logger = configurar_logger()

# Nombre canónico de la hoja dinámica
HOJA_DINAMICA = "TablaDinamica"

# Candidatos de nombre para la hoja de datos principal
# (se usa obtener_hoja_excel para no depender de un nombre exacto)
CANDIDATOS_HOJA_DATOS = (
    "Transac. Pend. Por Contabilizar",
    "Transacciones pendientes por contabilizar",
    "Transac.Pend.Por Contabilizar",
)


def crear_tabla_dinamica(origen_excel: str, hoja_dinamica: str = HOJA_DINAMICA) -> bool:
    """
    Crea una tabla dinámica de conteo en una hoja separada del Excel.
    Retorna True si se creó correctamente, False si hubo un error.

    CAMBIO: usa obtener_hoja_excel() para no depender del nombre exacto de la
    hoja de datos. Antes fallaba silenciosamente si el nombre no coincidía.
    """
    try:
        # Resolver nombre de hoja de forma robusta
        nombre_hoja = obtener_hoja_excel(origen_excel, CANDIDATOS_HOJA_DATOS)
    except ValueError as e:
        logger.error(
            "No se encontró la hoja de datos en '%s' para crear la tabla dinámica: %s",
            origen_excel, e,
        )
        return False

    try:
        df = pd.read_excel(origen_excel, sheet_name=nombre_hoja)
    except Exception as e:
        logger.error(
            "No se pudo leer la hoja '%s' de '%s': %s",
            nombre_hoja, origen_excel, e,
        )
        return False

    if df.empty:
        logger.warning(
            "La hoja '%s' de '%s' está vacía. No se creará la tabla dinámica.",
            nombre_hoja, origen_excel,
        )
        return False

    # Validar que existan suficientes columnas
    columnas_requeridas = [0, 4, 15, 5]  # índices usados abajo
    if len(df.columns) <= max(columnas_requeridas):
        logger.error(
            "El DataFrame tiene %d columnas pero se necesitan al menos %d. "
            "Verifique la estructura del Excel '%s'.",
            len(df.columns), max(columnas_requeridas) + 1, origen_excel,
        )
        return False

    indice_1 = df.columns[0]
    indice_2 = df.columns[4]
    indice_3 = df.columns[15]
    columna_valor = df.columns[5]

    try:
        pivot = pd.pivot_table(
            df,
            index=[indice_1, indice_2, indice_3],
            values=[columna_valor],
            aggfunc="count",
        )
    except Exception as e:
        logger.error("Error al crear la tabla dinámica: %s", e)
        return False

    try:
        with pd.ExcelWriter(
            origen_excel, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            pivot.to_excel(writer, sheet_name=hoja_dinamica)
    except PermissionError:
        logger.error(
            "Permiso denegado al escribir la tabla dinámica en '%s'. "
            "Cierre el archivo si lo tiene abierto en Excel.",
            origen_excel,
        )
        return False
    except Exception as e:
        logger.error("Error al escribir la tabla dinámica en '%s': %s", origen_excel, e)
        return False

    # Aplicar bordes
    try:
        wb = load_workbook(origen_excel)
        if hoja_dinamica not in wb.sheetnames:
            logger.error(
                "La hoja '%s' no se creó en '%s'. No se aplicarán bordes.",
                hoja_dinamica, origen_excel,
            )
            return False

        ws = wb[hoja_dinamica]
        borde = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row,
            min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = borde

        wb.save(origen_excel)
        logger.info(
            "Tabla dinámica creada con bordes en hoja '%s' de '%s'.",
            hoja_dinamica, origen_excel,
        )
        print(f"✅ Tabla dinámica creada en hoja '{hoja_dinamica}'.")
        return True

    except Exception as e:
        logger.error(
            "Error al aplicar bordes en la tabla dinámica de '%s': %s",
            origen_excel, e,
        )
        return False
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from .hojas_excel import obtener_hoja_excel
#Transacciones pendientes por contabilizar en el filtro_17-Julio-2025_06am

def crear_tabla_dinamica(origen_excel, hoja_dinamica='TablaDinamica'):
    # Leer la hoja principal de transacciones, tolerando nombres alternativos del archivo.
    hoja = obtener_hoja_excel(origen_excel, ("Transac. Pend. Por Contabilizar", "Transacciones pendientes por contabilizar"))
    df = pd.read_excel(origen_excel, sheet_name=hoja)

    # Usar índices de columnas para referenciar
    indice_1 = df.iloc[:, 0].name  # Columna 3
    indice_2 = df.iloc[:, 4].name
    indice_3 = df.iloc[:, 15].name  # Columna 4
    columna_valor = df.iloc[:, 5].name  # Columna 5

    # Crear tabla dinámica con conteo (count)
    pivot = pd.pivot_table(df,
                           index=[indice_1, indice_2, indice_3],
                           values=[columna_valor],
                           aggfunc='count')  # cuenta cuántas veces aparece

    # Escribir en hoja nueva
    with pd.ExcelWriter(origen_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        pivot.to_excel(writer, sheet_name=hoja_dinamica)
         # Ahora aplicar bordes con openpyxl

    wb = load_workbook(origen_excel)
    ws = wb[hoja_dinamica]

    # Crear un borde negro delgado
    borde = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Aplicar borde a todas las celdas con datos
    for row in ws.iter_rows(min_row=1,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column):
        for cell in row:
            cell.border = borde

    wb.save(origen_excel)
    print(f"✅ Tabla dinámica con bordes creada en hoja '{hoja_dinamica}' del archivo.")
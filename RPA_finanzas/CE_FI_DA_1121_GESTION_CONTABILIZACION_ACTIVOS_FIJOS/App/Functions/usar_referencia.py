import pandas as pd
from openpyxl import load_workbook
import shutil
from pathlib import Path
from .logs_config import configurar_logger
logger = configurar_logger()

def actualizar_pu_je(df_ref: pd.DataFrame, df_new: pd.DataFrame) -> None:

    df_ref_filtrado = df_ref[df_ref.iloc[:, 4].isin(["JE", "PU", "A5"])]
    ref_dict = dict(zip(df_ref_filtrado.iloc[:, 21], df_ref_filtrado.iloc[:, 0])) # Mapea orden -> responsable
    mask_tipos = df_new.iloc[:, 4].isin(["JE", "PU", "A5"])
    df_filtrado = df_new[mask_tipos]
    # Contar actualizaciones
    actualizaciones = 0
    for idx, row in df_filtrado.iterrows():
        orden = row.iloc[21]
        tipo = row.iloc[4]
        
        print(f"Evaluando {tipo}-{orden}")
        
        if pd.notna(orden) and orden in ref_dict:
            responsable = ref_dict[orden]
            valor_actual = df_new.at[idx, df_new.columns[0]]
            if pd.isna(valor_actual) or str(valor_actual).strip() == "":
                df_new.at[idx, df_new.columns[0]] = responsable
            actualizaciones += 1
            print(f"Actualizado: {orden} -> {responsable}")
        else:
            print(f"No encontrado: {orden}")
    print(f"Total actualizaciones: {actualizaciones}")

def actualizar_pv_ov(df_ref, df_new):
    # Crear diccionario de referencia para PV y OV
    df_ref_filtrado = df_ref[df_ref.iloc[:, 4].isin(["PV", "OV"])]
    # Crear clave compuesta: orden + cuenta
    claves_ref = df_ref_filtrado.iloc[:, 21].astype(str) + "_" + df_ref_filtrado.iloc[:, 15].astype(str)  # Columna P es la 15
    responsables_ref = df_ref_filtrado.iloc[:, 0]
    # Crear diccionario {clave_compuesta: responsable}
    ref_dict = dict(zip(claves_ref, responsables_ref))
    mask = df_new.iloc[:, 4].isin(["PV", "OV"])
    df_filtrado = df_new[mask]

    print(f"Registros PV/OV a evaluar: {len(df_filtrado)}")
    claves_nuevas = df_new.loc[mask, df_new.columns[21]].astype(str) + "_" + df_new.loc[mask, df_new.columns[15]].astype(str)
    # Mapear valores
    valores_mapeados = claves_nuevas.map(ref_dict)
    mask_coincidencias = mask & valores_mapeados.notna()
    mask_vacio = mask_coincidencias & (df_new.loc[mask_coincidencias, df_new.columns[0]].isna() | df_new.loc[mask_coincidencias, df_new.columns[0]].astype(str).str.strip().eq(""))
    df_new.loc[mask_vacio, df_new.columns[0]] = valores_mapeados[mask_vacio]
    print(f"Actualizaciones realizadas: {mask_coincidencias.sum()}")
   
def actualizar_ct(df_ref,df_new):
    # Filtrar registros CT en ambos DataFrames
    df_new_ct = df_new[df_new.iloc[:, 4] == "CT"]
    df_ref_ct = df_ref[df_ref.iloc[:, 4] == "CT"]


    df_ref_con_orden = df_ref_ct[df_ref_ct.iloc[:, 21].notna() & (df_ref_ct.iloc[:, 21] != "")]
    dict_con_orden = dict(zip(df_ref_con_orden.iloc[:, 21].astype(str), df_ref_con_orden.iloc[:, 0]))
      
    # Filtrar registros nuevos CT que tienen número de orden
    mask_nuevos_con_orden = (df_new.iloc[:, 4] == "CT") & df_new.iloc[:, 21].notna() & (df_new.iloc[:, 21] != "")
    
    # Actualizar registros con orden
    if mask_nuevos_con_orden.sum() > 0:
        valores_mapeados_orden = df_new.loc[mask_nuevos_con_orden, df_new.columns[21]].astype(str).map(dict_con_orden)
        mask_coincidencias_orden = mask_nuevos_con_orden & valores_mapeados_orden.notna()
        mask_vacio_orden = mask_coincidencias_orden & (df_new.loc[mask_coincidencias_orden, df_new.columns[0]].isna() | df_new.loc[mask_coincidencias_orden, df_new.columns[0]].astype(str).str.strip().eq(""))
        df_new.loc[mask_vacio_orden, df_new.columns[0]] = valores_mapeados_orden[mask_vacio_orden]
        print(f"Actualizaciones CT con orden: {mask_coincidencias_orden.sum()}")
    
    # CASO 2: Registros CT sin número de orden (columna 21 vacía o None)
    # Crear diccionario para registros sin número de orden usando explicación (columna 2)
    df_ref_sin_orden = df_ref_ct[df_ref_ct.iloc[:, 21].isna() | (df_ref_ct.iloc[:, 21] == "")]
    dict_sin_orden = dict(zip(df_ref_sin_orden.iloc[:, 2].astype(str), df_ref_sin_orden.iloc[:, 0]))
    
    # Filtrar registros nuevos CT que NO tienen número de orden
    mask_nuevos_sin_orden = (df_new.iloc[:, 4] == "CT") & (df_new.iloc[:, 21].isna() | (df_new.iloc[:, 21] == ""))
    
    # Actualizar registros sin orden usando explicación
    if mask_nuevos_sin_orden.sum() > 0:
        valores_mapeados_explicacion = df_new.loc[mask_nuevos_sin_orden, df_new.columns[2]].astype(str).map(dict_sin_orden)
        mask_coincidencias_explicacion = mask_nuevos_sin_orden & valores_mapeados_explicacion.notna()
        mask_vacio_explicacion = mask_coincidencias_explicacion & (df_new.loc[mask_coincidencias_explicacion, df_new.columns[0]].isna() | df_new.loc[mask_coincidencias_explicacion, df_new.columns[0]].astype(str).str.strip().eq(""))
        df_new.loc[mask_vacio_explicacion, df_new.columns[0]] = valores_mapeados_explicacion[mask_vacio_explicacion]
        print(f"Actualizaciones CT sin orden (por explicación): {mask_coincidencias_explicacion.sum()}")

def guardar_excel(excel_actual,df_new):
    wb = load_workbook(excel_actual)
    ws = wb.active
    if ws is None:
        logger.error("No se pudo acceder a la hoja activa del libro Excel")
        return
    for index, row in df_new.iterrows():
        excel_row = index + 2
        ws.cell(row=excel_row, column=1, value=row.iloc[0])
    wb.save(excel_actual)
    print(f"Archivo guardado en: {excel_actual}")
    
def usar_excel_referencia(excel_refencia, excel_actual):
    df_ref = pd.read_excel(excel_refencia)
    df_new = pd.read_excel(excel_actual).copy()
    
    actualizar_pu_je(df_ref, df_new)
    actualizar_ct(df_ref,df_new)
    actualizar_pv_ov(df_ref,df_new)

    guardar_excel(excel_actual, df_new)
    print(f"se actualizo correctamente el excel usando el excel de referencia: {excel_refencia}")
    logger.info(f"se actualizo correctamente el excel usando el excel de referencia: {excel_refencia}")
    return

def eliminar_informes(ruta_de_archivos, numeros_companias):
    # Convertir a Path para mejor manejo
    ruta_base = Path(ruta_de_archivos)
    if not ruta_base.exists():
        print(f"Error: La ruta '{ruta_base}' no existe")
        return {"error": f"Ruta no encontrada: {ruta_base}"}
    
    if not ruta_base.is_dir():
        print(f"Error: '{ruta_base}' no es un directorio")
        return {"error": f"No es un directorio: {ruta_base}"}
        # Convertir números de compañía a strings para comparación
    numeros_companias = [str(num) for num in numeros_companias]
     # Buscar carpetas que coincidan
    carpetas_encontradas = []
    carpetas_eliminadas = []
    errores = []
    
    for item in ruta_base.iterdir():
        if item.is_dir() and item.name in numeros_companias:
            carpetas_encontradas.append(item)
    
    if not carpetas_encontradas:
        print("No se encontraron carpetas que coincidan con los números de compañía")
        return {"carpetas_encontradas": 0, "carpetas_eliminadas": 0, "errores": []}
    
    
    
    # Eliminar carpetas
    print("Eliminando carpetas...")
    for carpeta in carpetas_encontradas:
        try:
            shutil.rmtree(carpeta)
            carpetas_eliminadas.append(carpeta.name)
        except Exception as e:
            error_msg = f"Error al eliminar {carpeta.name}: {str(e)}"
            errores.append(error_msg)
            print(f"  ✗ {error_msg}")
    
    # Resumen final
    print(f"  Carpetas encontradas: {len(carpetas_encontradas)}")
    print(f"  Carpetas eliminadas: {len(carpetas_eliminadas)}")
    print(f"  Errores: {len(errores)}")
    

